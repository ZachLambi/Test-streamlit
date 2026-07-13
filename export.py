"""
export.py — Mise en forme et génération de fichiers d'export
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Objectif : que l'export SOIT la version finale (en-têtes clairs, colonnes
ajustées, gel des volets) — pas un point de départ à remanier dans Excel
avant de pouvoir l'utiliser.

Deux formes de tableau, réutilisées à l'identique pour l'écran ET l'export
(impossible qu'ils divergent) :
  - mettre_en_forme_principal() : tableau principal, une colonne par année,
    Valeur et Variation annuelle (%) empilées en DEUX lignes par groupe
    (pas deux colonnes) — la ligne Variation suit immédiatement sa ligne
    Valeur. CAGR en colonnes de fin, rempli seulement sur la ligne Valeur.
  - mettre_en_forme_metrique() : tableau séparé à une seule métrique par
    année (Part de marché, Rang) — une ligne par groupe, une colonne par
    année, pas de logique d'empilement (n'a pas de sens pour ces métriques).
"""

from io import BytesIO
import pandas as pd

NOMS_COLONNES_AFFICHAGE = {
    "annee": "Année", "source": "Source", "flux": "Flux",
    "origine": "Origine", "type_ori": "Type origine",
    "destination": "Destination", "type_dest": "Type destination",
    "hs6": "Code HS6", "valeur": "Valeur",
    "variation_pct": "Variation ann. (%)",
    "cagr_5ans_pct": "CAGR (%)", "n_annees_reel": "Années (CAGR)",
    "part_marche_pct": "Part de marché (%)", "rang": "Rang",
}

_CLES_IDENTITE = ["flux", "origine", "destination", "hs6"]


def _convertir_noms(df: pd.DataFrame, noms_geo: dict[str, str]) -> pd.DataFrame:
    df = df.copy()
    df["origine"] = df["origine"].map(lambda c: noms_geo.get(c, c))
    df["destination"] = df["destination"].map(lambda c: noms_geo.get(c, c))
    return df


def _colonnes_annees(df: pd.DataFrame) -> list[str]:
    """Colonnes dont l'en-tête est une année brute (ex: '2019') — pas
    'Année', pas les autres colonnes."""
    return [c for c in df.columns if c.isdigit() and len(c) == 4]


def _formater_montant(x) -> str | None:
    """Formate un montant avec des espaces insécables comme séparateur de
    milliers (ex: 1356867 -> '1 356 867'), convention typographique
    française — U+00A0, garanti peu importe le navigateur/OS, contrairement
    au format Excel qui dépend des paramètres régionaux de qui l'ouvre."""
    if x is None or pd.isna(x):
        return None
    return f"{x:,.0f}".replace(",", "\u00a0")


def formater_pour_ecran(df_principal: pd.DataFrame) -> pd.DataFrame:
    """Copie du tableau principal (déjà mis en forme, valeurs NUMÉRIQUES)
    où les cellules de la ligne 'Valeur' deviennent du texte avec espace
    insécable — pour l'affichage à l'écran SEULEMENT. Les exports (Excel,
    CSV) continuent d'utiliser la version numérique d'origine, pas
    celle-ci — c'est la seule différence entre écran et export dans cette
    app, faite exprès (Excel garde de vrais nombres utilisables, l'écran
    garantit le séparateur que Streamlit ne peut pas offrir nativement).

    Toute la colonne (pas seulement la ligne Valeur) est convertie en texte
    homogène — une colonne qui mélange str (Valeur) et float (Variation)
    fait planter la sérialisation interne de Streamlit vers Arrow (elle
    s'en sort par un correctif automatique, mais avec une erreur bruyante
    dans les logs) — éviter le mélange de types est plus propre que de
    compter sur ce correctif."""
    if df_principal.empty or "Mesure" not in df_principal.columns:
        return df_principal
    df = df_principal.copy()
    colonnes_annees = _colonnes_annees(df)

    def _formater_cellule(x, mesure: str):
        if pd.isna(x):
            return None
        return _formater_montant(x) if mesure == "Valeur" else f"{x:.2f}"

    for col in colonnes_annees:
        df[col] = [
            _formater_cellule(x, mesure) for x, mesure in zip(df[col], df["Mesure"])
        ]
    return df


def mettre_en_forme_principal(
    df: pd.DataFrame, noms_geo: dict[str, str], avec_variation: bool
) -> pd.DataFrame:
    """
    Tableau principal : une colonne par année, Valeur et Variation annuelle
    (%) empilées en DEUX LIGNES par groupe (flux/origine/destination/hs6)
    plutôt qu'en colonnes séparées — la colonne 'Mesure' indique laquelle
    des deux. La ligne Variation suit toujours immédiatement sa ligne
    Valeur (ordre de construction, pas un tri après coup).

    Les montants de la ligne Valeur restent de VRAIS NOMBRES (pas du
    texte) — utilisables dans une formule Excel, triables normalement.
    L'export Excel leur applique un format d'affichage '#,##0' qui montre
    un séparateur de milliers selon les paramètres régionaux de l'Excel
    qui ouvre le fichier (espace pour un Excel en français, virgule pour
    un Excel en anglais — pas garanti, mais reste un vrai nombre). À
    l'écran (Streamlit), aucun séparateur n'est possible du tout — limite
    documentée de st.dataframe, pas un choix de ce module.

    CAGR (et son nombre d'années réel) en colonnes à la toute fin, remplies
    UNIQUEMENT sur la ligne Valeur (vides sur la ligne Variation) — c'est
    une métrique par SÉRIE, pas par année, ça n'a pas de sens de la répéter
    ou de l'associer à la ligne Variation.
    """
    if df.empty:
        return df

    df = _convertir_noms(df, noms_geo)
    cles = [c for c in _CLES_IDENTITE if c in df.columns]
    annees = sorted(df["annee"].dropna().unique())

    colonne_cagr = next((c for c in df.columns if c.startswith("cagr_") and c.endswith("ans_pct")), None)
    colonnes_serie = ([colonne_cagr] if colonne_cagr else []) + (
        ["n_annees_reel"] if "n_annees_reel" in df.columns else []
    )
    a_variation = avec_variation and "variation_pct" in df.columns

    lignes = []
    for cles_valeurs, sous in df.groupby(cles, dropna=False, sort=False):
        if not isinstance(cles_valeurs, tuple):
            cles_valeurs = (cles_valeurs,)
        sous = sous.set_index("annee")
        base_dict = dict(zip(cles, cles_valeurs))

        ligne_valeur = {**base_dict, "Mesure": "Valeur"}
        for annee in annees:
            ligne_valeur[str(annee)] = sous["valeur"].get(annee)
        for col_serie in colonnes_serie:
            if col_serie in sous.columns:
                valeurs = sous[col_serie].dropna()
                ligne_valeur[col_serie] = valeurs.iloc[0] if not valeurs.empty else None
            else:
                ligne_valeur[col_serie] = None
        lignes.append(ligne_valeur)

        if a_variation:
            ligne_variation = {**base_dict, "Mesure": "Variation annuelle (%)"}
            for annee in annees:
                ligne_variation[str(annee)] = sous["variation_pct"].get(annee)
            for col_serie in colonnes_serie:
                ligne_variation[col_serie] = None  # vide sur la ligne Variation, par design
            lignes.append(ligne_variation)

    resultat = pd.DataFrame(lignes)
    return resultat.rename(columns=NOMS_COLONNES_AFFICHAGE)


def mettre_en_forme_metrique(df: pd.DataFrame, colonne: str, noms_geo: dict[str, str]) -> pd.DataFrame:
    """
    Tableau séparé pour UNE métrique par année qui n'a pas de logique
    Valeur/Variation à empiler (Part de marché, Rang) — une ligne par
    groupe, une colonne par année contenant directement cette métrique.
    """
    if df.empty or colonne not in df.columns:
        return pd.DataFrame()

    df = _convertir_noms(df, noms_geo)
    cles = [c for c in _CLES_IDENTITE if c in df.columns]
    annees = sorted(df["annee"].dropna().unique())

    base = df[cles].drop_duplicates(subset=cles).reset_index(drop=True)
    for annee in annees:
        sous_annee = df[df["annee"] == annee]
        serie = sous_annee.set_index(cles)[colonne].rename(str(annee))
        base = base.merge(serie, on=cles, how="left")

    return base.rename(columns=NOMS_COLONNES_AFFICHAGE)


def exporter_excel(tables: dict[str, pd.DataFrame], note_unite: str) -> bytes:
    """Génère un .xlsx avec UNE FEUILLE PAR TABLE fournie (ex: "Extraction",
    "Part de marché", "Rang") — chaque table déjà mise en forme par
    mettre_en_forme_principal()/mettre_en_forme_metrique(). Tables vides
    ignorées. En-têtes en gras avec fond, colonnes ajustées, volets gelés
    sur chaque feuille. Retourne les bytes du fichier.

    Import openpyxl différé (dans la fonction, pas en haut du module) —
    évite un conflit bas niveau observé entre openpyxl et DuckDB quand les
    deux sont chargés dans le même contexte avant tout appel Excel réel."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nom_feuille, df_table in tables.items():
            if df_table is None or df_table.empty:
                continue
            nom_feuille_court = nom_feuille[:31]  # limite Excel
            df_table.to_excel(writer, sheet_name=nom_feuille_court, index=False, startrow=1)
            ws = writer.sheets[nom_feuille_court]

            ws.cell(row=1, column=1, value=f"Unités : {note_unite}").font = Font(italic=True, size=9)

            fond_entete = PatternFill(start_color="E8EAED", end_color="E8EAED", fill_type="solid")
            for col_idx, nom_col in enumerate(df_table.columns, start=1):
                cellule = ws.cell(row=2, column=col_idx)
                cellule.font = Font(bold=True)
                cellule.fill = fond_entete
                cellule.alignment = Alignment(horizontal="center")

            for col_idx, nom_col in enumerate(df_table.columns, start=1):
                lettre = get_column_letter(col_idx)
                largeur_contenu = df_table[nom_col].astype(str).str.len().max()
                largeur = max(len(str(nom_col)), int(largeur_contenu) if pd.notna(largeur_contenu) else 10) + 2
                ws.column_dimensions[lettre].width = min(largeur, 40)

            # Format numérique '#,##0' sur les colonnes années, SEULEMENT
            # pour les lignes "Valeur" (pas "Variation annuelle (%)", qui
            # partage les mêmes colonnes mais reste un pourcentage). Le
            # séparateur affiché dépend des paramètres régionaux de l'Excel
            # qui ouvre le fichier — voir mettre_en_forme_principal().
            colonnes_annees = _colonnes_annees(df_table)
            if "Mesure" in df_table.columns and colonnes_annees:
                idx_colonnes_annees = [df_table.columns.get_loc(c) + 1 for c in colonnes_annees]
                for i, valeur_mesure in enumerate(df_table["Mesure"]):
                    if valeur_mesure == "Valeur":
                        ligne_excel = i + 3  # +1 (1-index) +1 (note) +1 (en-tête)
                        for col_idx in idx_colonnes_annees:
                            ws.cell(ligne_excel, col_idx).number_format = "#,##0"

            ws.freeze_panes = "A3"

    return buffer.getvalue()


def exporter_csv(df_principal: pd.DataFrame) -> bytes:
    """CSV du tableau PRINCIPAL uniquement (un CSV = une seule table par
    nature — les tableaux Part de marché/Rang, s'il y en a, ne sont
    disponibles qu'en Excel, en feuilles séparées)."""
    return df_principal.to_csv(index=False).encode("utf-8-sig")  # BOM pour Excel FR